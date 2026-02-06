import warnings
from openpyxl import load_workbook

warnings.filterwarnings(
    "ignore",
    category=UserWarning,
    module="openpyxl"
)


def get_excel_cell_value(exc_file, exc_sheet, cell_number):
    wb = load_workbook(exc_file, data_only=True)
    ws = wb[exc_sheet]
    return ws[cell_number].value


def return_format(exc_file, exc_sheet, cell_number):

  FORMAT_MAP = {
    "1CT": 1,
    "2CT": 2,
    "3CT": 3
  }

  format_value = get_excel_cell_value(exc_file, exc_sheet, cell_number)
  format_divisor = FORMAT_MAP[format_value]

  return format_divisor


def sum_scraps(scrap_dic, key_stop, RAW_KEY, DIVIDED_KEY):
  raw_scrap = 0
  divided_scrap = 0
  for key in scrap_dic:

      if key == key_stop:
          break
        
      raw_scrap += scrap_dic[key][RAW_KEY]
      divided_scrap += scrap_dic[key][DIVIDED_KEY]

  return [raw_scrap, divided_scrap]
